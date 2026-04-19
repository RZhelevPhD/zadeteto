"""
meme_create_master_sheet.py

Generates (or regenerates) the Meme Machine master input sheet:
    0. MEME MACHINE - Copy/meme-master-sheet.xlsx

Five tabs: CONFIG, AVATARS, PRODUCTS, FORMULAS, QUEUE.
All columns follow the schema in directives/meme-create-master-sheet.md

Usage:
    python executions/meme_create_master_sheet.py
    python executions/meme_create_master_sheet.py --force
    python executions/meme_create_master_sheet.py --force --no-seed
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / "0. MEME MACHINE - Copy"
OUTPUT_FILE = OUTPUT_DIR / "meme-master-sheet.xlsx"

FONT_NAME = "Arial"
FONT_SIZE = 10
HEADER_FILL = PatternFill("solid", start_color="1F3A5F", end_color="1F3A5F")
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
WRAP = Alignment(wrap_text=True, vertical="top")

# ---------- schema definitions ----------

CONFIG_ROWS: list[tuple[str, object, str]] = [
    ("char_budget_multiplier", 1.25, "Multiplier for budget: sample_chars x multiplier"),
    ("funnel_L1_target_pct", 35, "Unaware"),
    ("funnel_L2_target_pct", 25, "Problem Aware"),
    ("funnel_L3_target_pct", 18, "Solution Aware"),
    ("funnel_L4_target_pct", 17, "Our Solution Aware"),
    ("funnel_L5_target_pct", 5, "Most Aware"),
    ("default_timezone", "Europe/Sofia", "For QUEUE post_time"),
    ("default_brand", "ZaDeteto", ""),
    ("default_brand_url", "https://zadeteto.bg", ""),
]

AVATARS_COLUMNS = [
    ("avatar_id", 10, "Primary key. Format A01, A02, ..."),
    ("avatar_name", 28, "Short Bulgarian label"),
    ("demographics", 38, "Age, location, income - one sentence"),
    ("role_context", 32, "Job, life stage"),
    ("core_pains", 55, "Key pains, separator '; '"),
    ("core_desires", 55, "Key desires, separator '; '"),
    ("fears", 40, "Fears that drive behavior"),
    ("daily_reality", 55, "Concrete daily scenes, separator '; '"),
    ("voice_style", 18, "Dropdown: sarcastic / warm / direct / skeptical / self-deprecating / earnest"),
    ("cultural_refs", 35, "Bulgarian cultural hooks, comma-separated"),
    ("notes", 30, ""),
]

PRODUCTS_COLUMNS = [
    ("product_id", 10, "Primary key. Format P01, P02, ..."),
    ("product_name", 32, ""),
    ("category", 22, ""),
    ("key_benefit", 50, "One sentence"),
    ("differentiation", 50, "vs competitors"),
    ("level_5_cta_bg", 55, "Bulgarian CTA text, pasted directly into L5 captions"),
    ("cta_url", 40, ""),
    ("notes", 30, ""),
]

FORMULAS_COLUMNS = [
    ("formula_id", 10, "Primary key. Format F01, F02, ..."),
    ("formula_name", 26, "Short label"),
    ("template_text_bg", 55, "Bulgarian template with [X], [Y], [Z] placeholders"),
    ("var_X_description", 32, "What fills the X slot"),
    ("var_Y_description", 32, "What fills the Y slot"),
    ("var_Z_description", 32, "What fills the Z slot"),
    ("sample_meme_text", 55, "Original source sample (any language)"),
    ("sample_char_count", 12, "Auto: =LEN(sample_meme_text)"),
    ("char_budget", 12, "Auto: =CEILING(sample_char_count * CONFIG multiplier, 1)"),
    ("theme_folder", 22, "Subfolder name under '0. Themes/'"),
    ("format_type", 18, "Dropdown: image+text / text_only / graph / list / speech_bubble / pie_chart / bar_chart"),
    ("source", 20, "Origin of formula"),
    ("notes", 30, ""),
]

QUEUE_COLUMNS = [
    ("queue_id", 10, "Primary key. Format Q001, Q002, ..."),
    ("post_date", 13, "YYYY-MM-DD"),
    ("post_time", 10, "HH:MM"),
    ("platform", 14, "Dropdown: Facebook / Instagram / LinkedIn / TikTok / Google"),
    ("product_id", 12, "FK -> PRODUCTS.product_id"),
    ("avatar_id", 12, "FK -> AVATARS.avatar_id"),
    ("formula_id", 12, "FK -> FORMULAS.formula_id"),
    ("awareness_level", 16, "Dropdown: 1-5"),
    ("status", 14, "Dropdown: pending / generated / approved / scheduled / posted / skipped"),
    ("notes", 40, ""),
]

# ---------- seed data ----------

AVATARS_SEED = [
    [
        "A01",
        "Заета работеща майка",
        "Жена 28-42, София/Пловдив, работеща, 1-2 деца",
        "Офис работа + пълен ден вкъщи",
        "Няма време; хроничната вина; чувството, че винаги изоставам; не знам кой специалист е добър",
        "Баланс; помощ без морал; някой да ми каже какво е нормално",
        "Че детето е 'не на ниво'; социална изолация; да сгреша избора",
        "Сутрешен хаос; вечерна умора; уикендите изчезват; WhatsApp групите не спират",
        "warm",
        "BG детски телевизионни предавания, баби в парка",
        "Primary persona",
    ],
    [
        "A02",
        "Родител на тинейджър",
        "Родител 40-55, цяла БГ, 1-2 тинейджъра",
        "Средна кариера, мислят за университет на детето",
        "Конфликти с тинейджъра; не разбирам какво става; страх от грешен път",
        "Близост без драма; да го/я спася от грешки; да видя, че е ОК",
        "Лоша среда; зависимости; отхвърляне от детето",
        "Вратата се затръшва; телефон на масата; мълчаливи вечери",
        "self-deprecating",
        "BG TV от 90-те, гимназиални спомени",
        "Secondary persona",
    ],
]

PRODUCTS_SEED = [
    [
        "P01",
        "ZaDeteto Directory Listing",
        "Directory",
        "Видимост в най-голямата БГ директория за детски специалисти",
        "Единствена с проверени отзиви + StoryBrand профили",
        "Регистрирай се за безплатен листинг на zadeteto.bg",
        "https://zadeteto.bg/register",
        "Free tier, lead-gen magnet",
    ],
    [
        "P02",
        "ZaDeteto Partner Package",
        "Premium Subscription",
        "Пълен брандиран профил + SEO + директен контакт с родители",
        "Премахва невидимостта от Google за соло специалисти",
        "Запазете място в Partner Package: https://zadeteto.bg/partners",
        "https://zadeteto.bg/partners",
        "Main revenue product",
    ],
]

FORMULAS_SEED = [
    [
        "F01",
        "Pringles Rick Astley",
        "Чухме, че [X] е толкова [Y], дори [Z] се отказва от тях.",
        "Нещо, което родителите мразят или е болезнена тема",
        "Прилагателно: скъп / изтощителен / безполезен / бавен",
        "Знаменитост или стереотипен персонаж, който би се отказал",
        "We heard that Pringles are so unhealthy, even Rick Astley is giving them up for 2024.",
        None,
        None,
        "Puns",
        "image+text",
        "Meme Maker Brian",
        "Seed formula",
    ],
    [
        "F02",
        "Line Graph Happiness",
        "Линейна графика с 2 линии: [X] намалява рязко, [Y] расте.",
        "Нещо, което родителите искат (напр. свободно време)",
        "Нещо, което всъщност получават (напр. уведомления във Viber)",
        "",
        "A line graph showing happiness declining and email stress increasing.",
        None,
        None,
        "Line Graphs",
        "graph",
        "Meme Maker Brian",
        "Two-variable only; leave Z blank",
    ],
    [
        "F03",
        "Dictionary Definition",
        "[X] (съществително): [Y]. Вижте също: [Z].",
        "Дума/термин от родителския свят",
        "Абсурдна, но истинска дефиниция",
        "Свързан термин за комичен ефект",
        "Monday (noun): the day that follows the weekend too quickly. See also: regret.",
        None,
        None,
        "Dictionary Definition",
        "text_only",
        "Meme Maker Brian",
        "Text-only; char budget важи за цялата дефиниция",
    ],
]

QUEUE_SEED = [
    [
        "Q001",
        "2026-04-22",
        "09:00",
        "Facebook",
        "P01",
        "A01",
        "F01",
        "2",
        "pending",
        "First test row - Tuesday morning drop",
    ],
    [
        "Q002",
        "2026-04-22",
        "18:30",
        "Instagram",
        "P01",
        "A01",
        "F02",
        "1",
        "pending",
        "Evening scroll slot",
    ],
    [
        "Q003",
        "2026-04-24",
        "10:00",
        "LinkedIn",
        "P02",
        "A02",
        "F03",
        "3",
        "pending",
        "",
    ],
]

# ---------- sheet-building helpers ----------

def write_header(ws: Worksheet, columns: list[tuple[str, int, str]]) -> None:
    for col_idx, (name, width, comment_text) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        if comment_text:
            cell.comment = Comment(comment_text, "meme-machine")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def write_rows(ws: Worksheet, rows: list[list[object]], start_row: int = 2) -> None:
    for r_offset, row in enumerate(rows):
        excel_row = start_row + r_offset
        for c_offset, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c_offset, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP


def add_list_validation(ws: Worksheet, col_letter: str, options: list[str], max_row: int = 500) -> None:
    """Attach an inline dropdown (Excel comma-separated list) to a column range."""
    quoted = '"' + ",".join(options) + '"'
    assert len(quoted) <= 255, (
        f"Inline list too long for Excel data validation ({len(quoted)} chars, max 255). "
        "Move options to a hidden sheet and use a range reference instead."
    )
    dv = DataValidation(type="list", formula1=quoted, allow_blank=True)
    dv.error = "Value must be one of: " + ", ".join(options)
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def add_range_validation(ws: Worksheet, col_letter: str, range_ref: str, max_row: int = 500) -> None:
    """Attach a dropdown sourced from a cross-sheet range (e.g. =PRODUCTS!$A$2:$A$200)."""
    dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
    dv.error = "Value must exist in the referenced table."
    dv.errorTitle = "Invalid reference"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


# ---------- tab builders ----------

def build_config(wb: Workbook, seed: bool) -> None:
    ws = wb.create_sheet("CONFIG")
    columns = [("key", 34, ""), ("value", 22, ""), ("notes", 55, "")]
    write_header(ws, columns)
    if seed:
        write_rows(ws, [list(r) for r in CONFIG_ROWS])


def build_avatars(wb: Workbook, seed: bool) -> None:
    ws = wb.create_sheet("AVATARS")
    write_header(ws, AVATARS_COLUMNS)
    if seed:
        write_rows(ws, AVATARS_SEED)
    add_list_validation(
        ws,
        "I",
        ["sarcastic", "warm", "direct", "skeptical", "self-deprecating", "earnest"],
    )


def build_products(wb: Workbook, seed: bool) -> None:
    ws = wb.create_sheet("PRODUCTS")
    write_header(ws, PRODUCTS_COLUMNS)
    if seed:
        write_rows(ws, PRODUCTS_SEED)


def build_formulas(wb: Workbook, seed: bool) -> None:
    ws = wb.create_sheet("FORMULAS")
    write_header(ws, FORMULAS_COLUMNS)

    # sample_char_count (col H) and char_budget (col I) are Excel formulas.
    # sample_meme_text = col G. char_budget_multiplier is looked up by KEY in CONFIG
    # (INDEX/MATCH on column A) so it survives row reorders in CONFIG.
    multiplier_lookup = 'INDEX(CONFIG!$B:$B,MATCH("char_budget_multiplier",CONFIG!$A:$A,0))'
    max_formula_row = 500
    for row in range(2, max_formula_row + 1):
        h = ws.cell(row=row, column=8)
        i = ws.cell(row=row, column=9)
        h.value = f'=IF(G{row}="","",LEN(G{row}))'
        i.value = f'=IF(H{row}="","",CEILING(H{row}*{multiplier_lookup},1))'
        h.font = BODY_FONT
        i.font = BODY_FONT
        h.alignment = Alignment(horizontal="right")
        i.alignment = Alignment(horizontal="right")

    if seed:
        # Write seed rows, but preserve the pre-filled formulas in H and I.
        for r_offset, row_data in enumerate(FORMULAS_SEED):
            excel_row = 2 + r_offset
            for c_offset, value in enumerate(row_data, start=1):
                if c_offset in (8, 9):  # skip H and I - formulas already set
                    continue
                cell = ws.cell(row=excel_row, column=c_offset, value=value)
                cell.font = BODY_FONT
                cell.alignment = WRAP

    add_list_validation(
        ws,
        "K",
        [
            "image+text",
            "text_only",
            "graph",
            "list",
            "speech_bubble",
            "pie_chart",
            "bar_chart",
        ],
    )


def build_queue(wb: Workbook, seed: bool) -> None:
    ws = wb.create_sheet("QUEUE")
    write_header(ws, QUEUE_COLUMNS)
    if seed:
        write_rows(ws, QUEUE_SEED)
    add_list_validation(ws, "D", ["Facebook", "Instagram", "LinkedIn", "TikTok", "Google"])
    add_list_validation(ws, "H", ["1", "2", "3", "4", "5"])
    add_list_validation(
        ws,
        "I",
        ["pending", "generated", "approved", "scheduled", "posted", "skipped"],
    )
    add_range_validation(ws, "E", "=PRODUCTS!$A$2:$A$200")
    add_range_validation(ws, "F", "=AVATARS!$A$2:$A$200")
    add_range_validation(ws, "G", "=FORMULAS!$A$2:$A$200")


# ---------- main ----------

def build_workbook(seed: bool) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    build_config(wb, seed)
    build_avatars(wb, seed)
    build_products(wb, seed)
    build_formulas(wb, seed)
    build_queue(wb, seed)
    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate the Meme Machine master input sheet.")
    parser.add_argument("--force", action="store_true", help="Overwrite existing file without asking.")
    parser.add_argument("--no-seed", action="store_true", help="Skip seed rows (headers only).")
    args = parser.parse_args()

    if not OUTPUT_DIR.exists():
        print(f"ERROR: output folder does not exist: {OUTPUT_DIR}", file=sys.stderr)
        return 2

    if OUTPUT_FILE.exists() and not args.force:
        print(
            f"ERROR: {OUTPUT_FILE.name} already exists. Pass --force to overwrite.",
            file=sys.stderr,
        )
        return 1

    wb = build_workbook(seed=not args.no_seed)
    wb.save(OUTPUT_FILE)
    size_kb = OUTPUT_FILE.stat().st_size / 1024
    print(f"OK: wrote {OUTPUT_FILE} ({size_kb:.1f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
